import os
from openpyxl import load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta

import pdb
'''
Customized for MSA output result, for different Insititude
Ministry/Statistics/year-end report/
'''
# WORKPLACE_PATH
# file_dir = r'C:\\Users\\Administrator\\Desktop\\msa_statics\\motor.xlsx'
# file_dir = r'C:\Users\Administrator\Desktop\全县市场主体2019-12-31（2020-01-16导出）.xlsx'
# output_dir = r'C:\\Users\\Administrator\\Desktop\\查询结果\\'
# edz_dir = r'D:\重装系统 0328 - 文档\_信息中心\_窗口查询\金钟管委会\金钟报表2019-4.xlsx'

# MBP_DIR
file_dir = r'/Users/amazm/Pprojects/TestLab/全县市场主体2019-12-31（2020-01-16导出）.xlsx'
edz_dir = r'/Users/amazm/Pprojects/TestLab/金钟报表2019-4.xlsx'


subject_attrs = ['个体', '农合', '内资', '私营', '外资']
subject_insts = ['individual', 'cooperative', 'domestic', 'private', 'forgein']
person_coefficient = [1.05, 6.5, 3, 3, 3]
theYearStart = datetime(2019, 1, 1)
theYearEnd = datetime(2019, 12, 31)
subjects = []

# JinZhong economic development zone
theAddress = '金钟'
spring = datetime(2019,3,31)
summer = datetime(2019,6,30)
autumn = datetime(2019,9,30)
winter = datetime(2019,12,31)



class Subject():

    def __init__(self):
        self.code = ''
        self.name = ''
        self.count = 0
        self.attr = ''
        self.address = ''
        self.person = 0
        self.regestied_time = ''

        self.fund_list = []

        # sum
        self.counts = 0
        self.people = 0
        self.fund_sum = 0

        # this year after theYearStart
        self.theFund_list = []
        self.theFund_sum = 0
        self.theCounts = 0
        self.thePeople = 0

        # the address
        self.theAddressFund_list = []
        self.theAddressFund_sum = 0


for i in range(5):
    subject_insts[i] = Subject()
    subject_insts[i].attr = subject_attrs[i]
    subjects.append(subject_insts[i])

company = Subject()


class Report():

    def __init__(self):
        pass

    def load_excel(self):
        wb = load_workbook(file_dir)
        ws = wb.active
        for i in range(2, ws.max_row + 1):
            name = ws[f'B{i}'].value
            attr = ws[f'G{i}'].value
            address = ws[f'M{i}'].value
            regestiedFund = ws[f'P{i}'].value
            regestiedTime = ws[f'AB{i}'].value

            if attr == subject_attrs[0]:
                subjects[0].fund_list.append(regestiedFund)
                if regestiedTime >= theYearStart:
                    subjects[0].theFund_list.append(regestiedFund)
                if theAddress in address:
                    subjects[0].theAddressFund_list.append(regestiedFund)

            elif attr == subject_attrs[1]:
                subjects[1].fund_list.append(regestiedFund)
                if regestiedTime >= theYearStart:
                    subjects[1].theFund_list.append(regestiedFund)
            elif attr == subject_attrs[2]:
                subjects[2].fund_list.append(regestiedFund)
                if regestiedTime >= theYearStart:
                    subjects[2].theFund_list.append(regestiedFund)
            elif attr == subject_attrs[3]:
                subjects[3].fund_list.append(regestiedFund)
                if regestiedTime >= theYearStart:
                    subjects[3].theFund_list.append(regestiedFund)
            elif attr == subject_attrs[4]:
                subjects[4].fund_list.append(regestiedFund)
                if regestiedTime >= theYearStart:
                    subjects[4].theFund_list.append(regestiedFund)

            else:
                print('stackoverflow ', attr, name,
                      regestiedFund, regestiedTime)

    def caculator(self):

        subjects_sum = 0
        the_subjects_sum = 0
        grow_speed = 0

        for i in range(5):
            subjects[i].counts = len(subjects[i].fund_list)
            subjects[i].fund_sum = sum(subjects[i].fund_list)
            subjects[i].people = round(
                subjects[i].counts * person_coefficient[i])
            subjects_sum += subjects[i].counts

            # the year after theYearStart
            subjects[i].theCounts = len(subjects[i].theFund_list)
            subjects[i].theFund_sum = sum(subjects[i].theFund_list)
            subjects[i].thePeople = round(
                subjects[i].theCounts * person_coefficient[i])
            the_subjects_sum += subjects[i].theCounts

        # grow speed
        grow_speed = round((subjects_sum /
                            (subjects_sum - the_subjects_sum) - 1) * 100, 2)

        # company
        for i in range(2, 5):
            company.counts += subjects[i].counts
            company.fund_sum += subjects[i].fund_sum
            company.people += subjects[i].people

            company.theCounts += subjects[i].theCounts
            company.theFund_sum += subjects[i].theFund_sum
            company.thePeople += subjects[i].thePeople

        # yearly report
        theYearEnd_str = theYearEnd.strftime(
            '%Y{y}%m{m}%d{d}').format(y='年', m='月', d='日')

        yearlyReport_str = [
            f'截止{theYearEnd_str}，全县市场主体总数为{subjects_sum}户，',
            f'其中企业{company.counts}户，注册资金{round(company.fund_sum, 4)}万元，',
            f'个体工商户{subjects[0].counts}户，注册资金{round(subjects[0].fund_sum, 4)}万元，',
            f'农民专业合作社{subjects[1].counts}户，注册资金{round(subjects[1].fund_sum, 4)}万元。',
            f'2019年全县新增企业{company.theCounts}户，注册资金{round(company.theFund_sum, 4)}万元，',
            f'新增个体工商户{subjects[0].theCounts}户，注册资金{round(subjects[0].theFund_sum, 4)}万元，',
            f'新增农民专业合作社{subjects[1].theCounts}户，注册金额{round(subjects[1].theFund_sum, 4)}万元。',
            f'市场主体发展同比增长{grow_speed}% 。'
        ]

        print(''.join(yearlyReport_str))


def main():
    R = Report()
    R.load_excel()
    R.caculator()


if __name__ == "__main__":
    main()
